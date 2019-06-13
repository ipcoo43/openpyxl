# 파이썬+엑셀: openpyxl 로 읽기/쓰기
# 두 엑셀 파일 읽어 하나로 합치기
# 범위 접근 - sheet['B2':'B7']
# 계산 및 결과 저장 - cell.value = val
# 셀에 스타일 지정 - Font, Alignment, Border, Side, Color, PatternFill
# 차트 그리기 : openpyxl.chart Reference, Series, BarChart : sheet.add_chart(chart, 'E1')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, Color, PatternFill
from openpyxl.chart import Reference, Series, BarChart


# 엑셀 파일 열기
df_samsung = pd.read_excel('./xlsx/2017년 광고비 - 삼성전자.xlsx')
df_samsung.set_index('date', inplace=True)

df_lg = pd.read_excel('./xlsx/2017년 광고비 - LG전자.xlsx')
df_lg.set_index('date', inplace=True)

df_merge = pd.DataFrame()
df_merge['samsung'] = df_samsung['total']
df_merge['lg'] = df_lg['total']

# 병합 파일 엑셀파일로 저장
df_merge.to_excel('./xlsx/merged_102_01.xlsx')

# 계산 및 계산 결과 저장
# 엑셀 파일 열고 시트 얻기
wb = openpyxl.load_workbook('./xlsx/merged_102_01.xlsx')
sheet = wb.active
sheet

# 삼성전자 컬럼
samsung_month_cost = [row[0].value for row in sheet['B2':'B13']]
samsung_month_cost
# 합계
sum(samsung_month_cost)
sheet['B14'].value = sum(samsung_month_cost)
sheet['B14'].value
sheet['A14'].value = '합계'

# 합계 계산
sheet['B14'].value = sum([row[0].value for row in sheet['B2':'B13']])
sheet['C14'].value = sum([row[0].value for row in sheet['C2':'C13']])
wb.save('./xlsx/merged_102_02.xlsx')

# 수식 넣기
# 합계 계산
sheet['B14'].value = '=SUM(B2:B13)'
sheet['C14'].value = '=SUM(C2:C13)'
wb.save('./xlsx/merged_102_03.xlsx')

# 셀에 스타일 지정
# Font: 'D2Coding', size 15, bold
font_15 = Font(name='D2Coding', size=15, bold=True)

# Alignment : 가로 세로, 가운데 정렬
align_center = Alignment(horizontal='center', vertical='center')
align_vcenter = Alignment(vertical='center')

# Border : 테두리 지정
border_thin = Border(
    left = Side(style='thin'), 
    right = Side(style='thin'),
    top  = Side(style='thin'),
    bottom = Side(style='thin')
)

# PatternFill: 셀 색상 지정
fill_orange = PatternFill(patternType='solid', fgColor=Color('FFC000'))
fill_lightgrey = PatternFill(patternType='solid', fgColor=Color('D3D3D3'))

cell_sum = sheet['A14'] # 합계 제목 셀
cell_sum.font = font_15
cell_sum.alignment = align_center
cell_sum.border = border_thin
cell_sum.fill = fill_orange
wb.save('./xlsx/merged_102_04.xlsx')

# 범위에 스타일 지정
sheet['B2:c14']   # 범위(range )
for row in sheet['B2:c14']:
    for cell in row:
        cell.border = border_thin
        cell.number_format = '0.00'

for row in sheet['B14:C14']:
    for cell in row:
        cell.alignment = align_vcenter
        cell.fill = fill_orange
wb.save('./xlsx/merged_102_05.xlsx')

# 차트 추가 하기
chart = BarChart()
chart.title = '2017년 월별 광고비 (억원)'

values = Reference(sheet, range_string='Sheet1!B1:B13')
series = Series(values, title='삼성전자')
chart.append(series)

values = Reference(sheet, range_string='Sheet1!C1:C13')
series = Series(values, title='LG전자')
chart.append(series)

sheet.add_chart(chart,'E1')

wb.save('./xlsx/merged_102_06.xlsx')