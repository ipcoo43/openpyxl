#엑셀 필터링과 소트 - 불린 인덱싱과 df.sort_values()
#엑셀 VLOOKUP - DataFrame 컬럼 복사만으로 가능 (더 복잡한 pandas 의 join 활용)
#엑셀 피벗 - pd.pivot_table()
# 필터링과 소트
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, Color, PatternFill

df_factor = pd.read_excel('./xlsx/종목별 팩터 데이터.xlsx', dtype={'종목코드':str})
df_factor = df_factor.set_index('종목코드')
print(df_factor)

# 필터링
print(df_factor[df_factor['PER'] <= 10])

# 배당수익률 내림차순 정렬
df = df_factor[df_factor['PER'] <=10].sort_values('배당수익률',ascending=False)
print(df)

# 조건에 해당하는 상위 3개 종목
print(df[['종목명']].head(3))

# vlookup
df_interst_stock = pd.read_excel('./xlsx/관심종목.xlsx', dtype={'종목코드':str})
df_interst_stock = df_interst_stock.set_index('종목코드')
print(df_interst_stock)
print(df_factor)

# 컬럼 병합 : Vlookup 기능은 pandas의 join등으로 가능하나 인덱스를 활용하면 간편하게 컬럼 복사만으로 가능 함
# 개별 컬럼 복사(없으면 생성됨)
df_interst_stock['PEF'] = df_factor['PER']
print(df_interst_stock)

df_interst_stock[['PER','PBR','배당수익률']]  = df_factor[['PER','PBR','배당수익률']]
print(df_interst_stock)

df_interst_stock = df_interst_stock.dropna()
print(df_interst_stock)

# 엑셀 파일로 저장
# 병합 처리 결과를 포맷팅하여 엑셀로 저장

# Border: 테두리 지정
border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'), 
    top=Side(style='thin'), bottom=Side(style='thin')
)

# PatternFill: 셀 색상 지정
fill_orange = PatternFill(patternType='solid', fgColor=Color('FFC000'))
fill_lightgrey = PatternFill(patternType='solid', fgColor=Color('D3D3D3'))

writer = pd.ExcelWriter('./xlsx/과심팩터병합_103.xlsx', engine='openpyxl')
df_interst_stock.to_excel(writer, sheet_name='Sheet1')

wb = writer.book
sheet = writer.sheets['Sheet1']

# 컬럼 너비 지정
for col in list('BCDE'):
    sheet.column_dimensions[col].width = 14

#  테두리, 숫자 포맷 지정
for row in sheet['B2:E5']:
    for cell in row:
        cell.border = border_thin
        cell.number_format = '0.00'

# 헤더 색상 지정
for row in sheet['A1:E1']:
    for cell in row:
        cell.fill = fill_lightgrey
writer.save()

# 피벗 테이블
df = pd.read_excel('./xlsx/광고비 데이터 3사.xlsx', dtype={'종목코드':str})
table = pd.pivot_table(df,values='total', index='date', columns='name')
print(table)
